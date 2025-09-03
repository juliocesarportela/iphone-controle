import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analyze_excel_file(file_path):
    """Analisa a planilha Excel e extrai toda a estrutura e lógica de negócio"""
    
    print("=" * 60)
    print("ANÁLISE COMPLETA DA PLANILHA EXCEL")
    print("=" * 60)
    
    # Carrega o workbook
    wb = load_workbook(file_path, data_only=True)
    
    print(f"\n📊 PLANILHAS ENCONTRADAS:")
    for sheet_name in wb.sheetnames:
        print(f"   - {sheet_name}")
    
    # Analisa cada planilha
    for sheet_name in wb.sheetnames:
        print(f"\n" + "="*50)
        print(f"ANÁLISE DA PLANILHA: {sheet_name}")
        print("="*50)
        
        ws = wb[sheet_name]
        
        # Dimensões da planilha
        print(f"\n📏 DIMENSÕES:")
        print(f"   - Linhas: {ws.max_row}")
        print(f"   - Colunas: {ws.max_column}")
        
        # Extrai dados das primeiras 20 linhas para análise
        print(f"\n📋 ESTRUTURA DOS DADOS (primeiras 20 linhas):")
        print("-" * 80)
        
        for row in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):  # Primeiras 10 colunas
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    row_data.append(f"Col{col}: {value}")
            
            if row_data:
                print(f"Linha {row:2d}: {' | '.join(row_data)}")
        
        # Procura por fórmulas importantes
        print(f"\n🧮 FÓRMULAS ENCONTRADAS:")
        print("-" * 50)
        
        formula_count = 0
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if hasattr(cell, 'formula') and cell.formula:
                    print(f"   Célula {cell.coordinate}: {cell.formula}")
                    formula_count += 1
                    if formula_count > 20:  # Limita a 20 fórmulas para não sobrecarregar
                        print("   ... (mais fórmulas encontradas)")
                        break
            if formula_count > 20:
                break
        
        if formula_count == 0:
            print("   Nenhuma fórmula encontrada (valores já calculados)")
        
        # Procura por cabeçalhos/campos importantes
        print(f"\n🏷️ POSSÍVEIS CAMPOS/CABEÇALHOS:")
        print("-" * 50)
        
        headers = []
        for row in range(1, min(6, ws.max_row + 1)):  # Primeiras 5 linhas
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if any(keyword in cell.value.lower() for keyword in 
                          ['modelo', 'iphone', 'custo', 'valor', 'frete', 'cambio', 
                           'taxa', 'peso', 'grade', 'capacidade', 'quantidade']):
                        headers.append(f"   {cell.coordinate}: {cell.value}")
        
        for header in set(headers):  # Remove duplicatas
            print(header)
    
    print(f"\n" + "="*60)
    print("ANÁLISE CONCLUÍDA")
    print("="*60)

if __name__ == "__main__":
    file_path = "Modelo_Custos_EUA_PY_GSheets.xlsx"
    try:
        analyze_excel_file(file_path)
    except Exception as e:
        print(f"Erro ao analisar a planilha: {e}")
        print("Verifique se o arquivo existe e está no formato correto.")
