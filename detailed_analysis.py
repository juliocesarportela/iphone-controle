import pandas as pd
import openpyxl
from openpyxl import load_workbook

def detailed_excel_analysis(file_path):
    """Análise detalhada da planilha Excel com todas as colunas"""
    
    print("=" * 80)
    print("ANÁLISE DETALHADA DA PLANILHA EXCEL - TODAS AS COLUNAS")
    print("=" * 80)
    
    # Carrega usando pandas para análise mais detalhada
    try:
        df = pd.read_excel(file_path, sheet_name='CUSTOS')
        
        print(f"\n📊 INFORMAÇÕES GERAIS:")
        print(f"   - Linhas de dados: {len(df)}")
        print(f"   - Colunas: {len(df.columns)}")
        
        print(f"\n🏷️ TODAS AS COLUNAS IDENTIFICADAS:")
        print("-" * 60)
        for i, col in enumerate(df.columns, 1):
            print(f"   Col {i:2d} ({chr(64+i)}): {col}")
        
        print(f"\n📋 PRIMEIRAS 5 LINHAS DE DADOS:")
        print("-" * 80)
        print(df.head().to_string())
        
        print(f"\n📊 ESTATÍSTICAS DOS CAMPOS NUMÉRICOS:")
        print("-" * 60)
        numeric_cols = df.select_dtypes(include=['number']).columns
        for col in numeric_cols:
            print(f"\n{col}:")
            print(f"   Min: {df[col].min():.2f}")
            print(f"   Max: {df[col].max():.2f}")
            print(f"   Média: {df[col].mean():.2f}")
        
        print(f"\n📱 MODELOS DE IPHONE ENCONTRADOS:")
        print("-" * 40)
        if 'MODELO' in df.columns:
            modelos = df['MODELO'].value_counts()
            for modelo, count in modelos.items():
                print(f"   {modelo}: {count} unidades")
        
        print(f"\n💾 CAPACIDADES ENCONTRADAS:")
        print("-" * 30)
        if 'GB' in df.columns:
            capacidades = df['GB'].value_counts()
            for cap, count in capacidades.items():
                print(f"   {cap}GB: {count} unidades")
        
        print(f"\n🏆 GRADES ENCONTRADAS:")
        print("-" * 25)
        if 'GRADE' in df.columns:
            grades = df['GRADE'].value_counts()
            for grade, count in grades.items():
                print(f"   {grade}: {count} unidades")
        
        # Análise de valores típicos
        print(f"\n💰 ANÁLISE DE VALORES TÍPICOS:")
        print("-" * 40)
        
        valor_cols = ['VALOR EUA $', 'TAXA ADM $', 'FRETE EUA $', 'POL EUA $']
        for col in valor_cols:
            if col in df.columns:
                valores_unicos = df[col].unique()
                if len(valores_unicos) <= 10:
                    print(f"   {col}: {list(valores_unicos)}")
                else:
                    print(f"   {col}: Min={df[col].min():.2f}, Max={df[col].max():.2f}")
        
        # Verifica se há fórmulas ou padrões de cálculo
        print(f"\n🧮 VERIFICAÇÃO DE CÁLCULOS:")
        print("-" * 40)
        
        # Tenta identificar a fórmula do custo EUA
        if all(col in df.columns for col in ['VALOR EUA $', 'TAXA ADM $', 'FRETE EUA $', 'POL EUA $', 'CUSTO EUA +0,5%']):
            df_sample = df.head(10)
            print("   Verificando fórmula do CUSTO EUA +0,5%:")
            for idx, row in df_sample.iterrows():
                valor_eua = row['VALOR EUA $']
                taxa_adm = row['TAXA ADM $']
                frete_eua = row['FRETE EUA $']
                pol_eua = row['POL EUA $']
                custo_calculado = row['CUSTO EUA +0,5%']
                
                # Testa diferentes fórmulas
                formula1 = (valor_eua + taxa_adm + frete_eua + pol_eua) * 1.005
                formula2 = (valor_eua + taxa_adm + frete_eua + pol_eua) * 1.05
                formula3 = valor_eua + taxa_adm + frete_eua + pol_eua + (valor_eua * 0.005)
                
                print(f"   Linha {idx+2}: Real={custo_calculado:.2f}, Fórmula1={formula1:.2f}, Diff={abs(custo_calculado-formula1):.4f}")
        
        print(f"\n" + "="*80)
        print("ANÁLISE DETALHADA CONCLUÍDA")
        print("="*80)
        
    except Exception as e:
        print(f"Erro na análise: {e}")
        # Fallback para openpyxl
        print("Tentando com openpyxl...")
        wb = load_workbook(file_path, data_only=True)
        ws = wb['CUSTOS']
        
        print(f"\n📋 CABEÇALHOS (Linha 1):")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                print(f"   Col {col} ({chr(64+col)}): {cell.value}")

if __name__ == "__main__":
    file_path = "Modelo_Custos_EUA_PY_GSheets.xlsx"
    detailed_excel_analysis(file_path)
